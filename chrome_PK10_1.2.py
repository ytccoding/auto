from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#1.2修改單一投注數,增進效能,修改確認投注方式
def submitCheck():
    period = []
    sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = test_web.webPage()[i].text
    sheet_money["C"+str(len(sheet_money["B"]))].value = Account               
    period.append(test_web.K3_r("input[type=text]" ,6 , max_Money = "1")) #投注金額

    try:
        period.insert(0 ,test_web.timeTitle()) #期號
    except:
        period.insert(0 ,"秒秒平常沒有期號") #期號

    submitCheck = True
    while(submitCheck):
        test_web.elementClick("div[class='checkedListCon'] a[class='betBtn']" ,6)
        #sleep(2)
        if test_web.radioWord() == "YES":#秒秒彩
            sleep(10)
            submitCheck = False
        elif test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
            #sleep(2)
            if test_web.submitCheckOK() != "NG":
                test_web.elementClick("//span[.='确定']" ,8)
                submitCheck = False
           
    #sleep(10)
   
    sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
    sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
    sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號
    
    sheet_row = len(sheet_money["B"]) #投注金額填表
    for k in range(len(period[1])):
        sheet_money.cell(row = sheet_row ,column = k + 7).value = period[1][k]
    wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_賽車" + "投注金額.xlsx")
    
    
test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]

print("所有快樂彩(賽車)全餐投注")
wb = load_workbook("快樂彩投注用.xlsx")
sheet = wb["賽車"] # 獲取一張表
testNumber = input("測試站點序號:").strip()
for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        Account = str(sheet["F" + str(i)].value).strip()
        Password = str(sheet["G" + str(i)].value).strip()
        webPageSelect = str(sheet["I" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url) #目標網址
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']",6))
sleep(5)
test_web.webDriver.get(url) #目標網址

if test_web.webDriver.current_url != url:
    input("此彩種未開放或URL有誤,請檢查,按enter離開。")

period = []

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["快樂彩"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

ballBranchList = ["复式","和值","定位胆"]
ballBranchList2 = ["单式"]
ballList1 = ["定位胆"]
ballList2 = ["猜前五"]
ballList3 = ["猜前四"]
ballList4 = ["猜前三"]
ballList5 = ["猜前二"]
ballList6 = ["猜冠军","冠亚和","龙虎斗"]
ballList7 = ["大小单双"]
boxList = []
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH

for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
   if  webPageSelect != "1":
       test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁
       
   for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
       test_web.webPlayClick(j)
       for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
           test_web.webPlayBranchClick(k)
           if test_web.webPlay()[j].text in ballList6: #"猜冠军","冠亚和","龙虎斗"
               if test_web.webPlayBranch()[k].text in ballBranchList:
                   for m in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                       for n in range(len(test_web.webBall())):
                           test_web.webBallClick(n ,m)
                           test_web.elementClick("betBtn" ,1)
                   submitCheck()
           elif test_web.webPlay()[j].text in ballList7:#大小单双
               if test_web.webPlayBranch()[k].text in ballBranchList:
                   for m in range(len(test_web.elements("div[class='buyNumber fix dsds']" ,6))):
                       for n in range(len(test_web.webBallDsds(m))):
                           test_web.webBallDsdsClick(n ,m)
                           test_web.elementClick("betBtn" ,1)
                   submitCheck()                   
           elif test_web.webPlay()[j].text in ballList1: #定位胆
               clickList = [[[0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9]],\

                            [[5,6,7,8,9],\
			     [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4]],\
                             
                            [[0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4]],\
                             
                            [[5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9]]]
               if test_web.webPlayBranch()[k].text in ballBranchList:
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
           elif test_web.webPlay()[j].text in ballList2:#猜前五
               clickList = [[[0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9]],\
                             
                            [[5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4]],\
                             
                            [[0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4]],\
                             
                            [[5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9]]]
               if test_web.webPlayBranch()[k].text in ballBranchList:
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList2:
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01 02 03 04 05") #單式投注內容
                   sleep(2)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
           elif test_web.webPlay()[j].text in ballList3:#猜前四
               clickList = [[[0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9]],\
                             
                            [[5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4]],\
                             
                            [[0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4],\
                             [0,1,2,3,4]],\
                             
                            [[5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9],\
                             [5,6,7,8,9]]]
               if test_web.webPlayBranch()[k].text in ballBranchList:
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList2:
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01 02 03 04") #單式投注內容
                   sleep(2)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
           elif test_web.webPlay()[j].text in ballList4:#猜前三
               clickList =  [[[0,1,2,3,4],\
                              [0,1,2,3,4],\
                              [5,6,7,8,9]],\
                             
                             [[5,6,7,8,9],\
                              [5,6,7,8,9],\
                              [0,1,2,3,4]],\
                             
                             [[0,1,2,3,4],\
                              [0,1,2,3,4],\
                              [0,1,2,3,4]],\
                             
                             [[5,6,7,8,9],\
                              [5,6,7,8,9],\
                              [5,6,7,8,9]]]
               if test_web.webPlayBranch()[k].text in ballBranchList:
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList2:
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01 02 03") #單式投注內容
                   sleep(2)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
           elif test_web.webPlay()[j].text in ballList5:#猜前二
               clickList = [[[0,1,2,3,4],\
                             [5,6,7,8,9]],\
                             
                            [[5,6,7,8,9],\
                             [0,1,2,3,4]],\
                             
                            [[0,1,2,3,4],\
                             [0,1,2,3,4]],\
                             
                            [[5,6,7,8,9],\
                             [5,6,7,8,9]]]
               if test_web.webPlayBranch()[k].text in ballBranchList:
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList2:
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01 02") #單式投注內容
                   sleep(2)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
           else:
               print(test_web.webPlay()[j].text + "_" +test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")
           



sleep(600)#等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_賽車" + "投注金額.xlsx")

if len(error) != 1:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_賽車_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()
