# -*- coding: utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#1.2修改單一投注數,增進效能,修改確認投注方式,判斷彩種是否開放,等待時間,登入時間,帳號分開

def submitCheck():
    period = []
    sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = test_web.webPage()[i].text
    sheet_money["C"+str(len(sheet_money["B"]))].value = Account
    period.append(test_web.KL8("order_type" ,2)) #投注金額

    try:
        period.insert(0 ,test_web.timeTitle()) #期號
    except:
        period.insert(0 ,"秒秒彩平常沒有期號") #期號

    submitCheck = True
    while(submitCheck):
        test_web.elementClick("div[class='checkedListCon'] a[class='betBtn']" ,6)
        #sleep(2)
        if test_web.radioWord() != "NO":
            sleep(10)
            submitCheck = False
        if test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
            #sleep(2)
            if test_web.submitCheckOK() != "NG":
                test_web.elementClick("//span[.='确定']" ,8)
                submitCheck = False
           
    #sleep(10)
   
    sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
    sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
    sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號
    
    sheet_row = len(sheet_money["B"]) #投注金額填表
    for k in range(len(period[1])):
        sheet_money.cell(row = sheet_row ,column = k + 7).value = period[1][k]
    wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_11選5" + "投注金額.xlsx")

print("11選5玩法投注")
testNumber = input("測試站點序號:").strip()
accountNumber = input("測試帳號序號:").strip()
    
test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]

wb = load_workbook("11選5投注用.xlsx")
sheet = wb["11選5"] # 獲取一張表
wbAccount = load_workbook("前台帳號.xlsx")
sheetAccount = wbAccount["帳號"] # 獲取一張表

for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        webPageSelect = str(sheet["G" + str(i)].value).strip()
        if str(sheet["F" + str(i)].value).strip() == "None":
            waitSec = 600
        else:
            waitSec = int(str(sheet["F" + str(i)].value).strip())

for i in range(1,len(sheetAccount["B"])+1):
    if str(sheetAccount["B" + str(i)].value).strip() == str(accountNumber):
        Account = str(sheetAccount["D" + str(i)].value).strip()
        Password = str(sheetAccount["E" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url) #目標網址
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']",6))
#sleep(5)
timeCount = 0
while(test_web.webDriver.current_url != url):
    sleep(1)
    timeCount = timeCount + 1
    if timeCount >= 30:
        break
test_web.webDriver.get(url) #目標網址

html_source = test_web.webDriver.page_source
if test_web.webDriver.current_url != url:
    input("此URL有誤,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "您所访问的彩种不存在，即将返回购彩大厅" in html_source:
    input("此彩種未開放,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "Unexpected token u in JSON at position 0" in html_source:
    input("此彩種未獎金模板錯誤或是HOST錯誤,請檢查,按enter離開。")
    test_web.webDriver.quit() 

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["快樂彩"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

ballBranchList1 = ["和值","头尾龙虎","前二龙虎","后二龙虎","牛牛","定单双","前三和值","中三和值","后三和值","头尾和值","前二和值","后二和值"]
ballBranchList3 = ["一中一","二中二","三中三","四中四","五中五"]
ballBranchList4 = ["六中五","七中五","八中五"]
ballBranchList5 = ["复式","前三直选复式","前二直选复式",]
ballBranchList6 = ["和尾值","猜中位","猜必不出","前三组选复式","前二组选复式","前三一码不定位"]
ballBranchList7 = ["前三直选单式","前三组选单式","前二直选单式","前二组选单式"]
ballBranchList8 = ["前三组选胆拖","前二组选胆拖"]

ballList1 = ["任选"]
ballList2 = ["定位胆"]
ballList3 = ["三码"]
ballList4 = ["二码"]
boxList1 = ["和值","不定位"]
boxList2 = ["龙虎斗"]
boxList3 = ["趣味型"]
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH
for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
   if  webPageSelect != "1":
       test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁
       
   for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
       test_web.webPlayClick(j)
       for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
           test_web.webPlayBranchClick(k)
           if test_web.webPlay()[j].text in boxList1: #和值,"不定位"
               if test_web.webPlayBranch()[k].text in ballBranchList6:#和尾值,"前三一码不定位"
                   for m in range(len(test_web.webBall())):
                       test_web.webBallClick(m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList1:#和值
                   for m in range(len(test_web.elements("div[class='syx5CheckNum fix'] ins" ,6))):
                       test_web.elementsClickOne("div[class='syx5CheckNum fix'] ins" ,6 ,m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")
                   
           elif test_web.webPlay()[j].text in boxList2: #龙虎斗
               if test_web.webPlayBranch()[k].text in ballBranchList1:#头尾龙虎","前二龙虎","后二龙虎
                   for m in range(len(test_web.elements("div[class='buyNumber fix'] ins" ,6))):
                       test_web.elementsClickOne("div[class='buyNumber fix'] ins" ,6 ,m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

           elif test_web.webPlay()[j].text in ballList1: #任选
               if k <= 7:
                   m = 0
               elif k >= 8 and  k <= 15:
                   m = 1
               elif k > 15:
                   m = 2
               if test_web.webPlayBranch()[k].text in ballBranchList3 and test_web.webPlayBranchTitle()[m].text == "任选复式":#"一中一","二中二","三中三","四中四","五中五"
                   clickList = [[[0,1,2,3,4]],\
                             
                                [[5,6,7,8,9,10]],\
                             
                                [[0,1,2,3,9,10]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList3 and test_web.webPlayBranchTitle()[m].text == "任选单式":#單式"一中一","二中二","三中三","四中四","五中五""六中五","七中五","八中五"
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01,01 02,01 02 03,01 02 03 04,01 02 03 04 05") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList4 and test_web.webPlayBranchTitle()[m].text == "任选复式":#"六中五","七中五","八中五"
                   clickList = [[[0,1,2,3,4,5,6,7]],\
                             
                                [[3,4,5,6,7,8,9,10]],\
                             
                                [[0,1,2,3,4,5,9,10]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList4 and test_web.webPlayBranchTitle()[m].text == "任选单式":#單式"六中五","七中五","八中五"
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01 02 03 04 05 06,01 02 03 04 05 06 07,01 02 03 04 05 06 07 08") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList3 and test_web.webPlayBranchTitle()[m].text == "任选胆拖":#"二中二","三中三","四中四","五中五"
                   clickList = [[[0],\
                                 [2,3,4,5,6,7,8,9,10]],\
                             
                                [[2],\
                                 [4,5,6,7,8,9,10,0,1]],\
                             
                                [[4],\
                                 [6,7,8,9,10,0,1,2,3]],\
                             
                                [[6],\
                                 [8,9,10,0,1,2,3,4,5]],\

                                [[8],\
                                 [10,0,1,2,3,4,5,6,7]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList4 and test_web.webPlayBranchTitle()[m].text == "任选胆拖":#"六中五","七中五","八中五"
                   clickList = [[[1],\
                                 [2,3,4,5,6,7,8,9,10]],\
                             
                                [[3],\
                                 [4,5,6,7,8,9,10,0,1]],\
                             
                                [[5],\
                                 [6,7,8,9,10,0,1,2,3]],\
                             
                                [[7],\
                                 [8,9,10,0,1,2,3,4,5]],\

                                [[9],\
                                 [10,0,1,2,3,4,5,6,7]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList7:#任选單式
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "112") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

           elif test_web.webPlay()[j].text in ballList2: #定位胆
               if test_web.webPlayBranch()[k].text in ballBranchList5:#复式
                   clickList = [[[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [0,1,2,3,4]],\
                             
                                [[5,6,7,8,9,10],\
                                 [5,6,7,8,9,10],\
                                 [5,6,7,8,9,10]],\
                             
                                [[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [5,6,7,8,9,10]],\
                             
                                [[5,6,7,8,9,10],\
                                 [5,6,7,8,9,10],\
                                 [0,1,2,3,4]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

           elif test_web.webPlay()[j].text in boxList3: #趣味型
               if test_web.webPlayBranch()[k].text in ballBranchList1:#"牛牛","定单双"
                   for m in range(len(test_web.elements("div[class='buyNumber fix'] ins" ,6))):
                       test_web.elementsClickOne("div[class='buyNumber fix'] ins" ,6 ,m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList6:#"猜中位","猜必不出"
                   for m in range(len(test_web.webBall())):
                       test_web.webBallClick(m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

           elif test_web.webPlay()[j].text in ballList3: #三码
               if test_web.webPlayBranch()[k].text in ballBranchList5:#前三直选复式
                   clickList = [[[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [0,1,2,3,4]],\
                             
                                [[5,6,7,8,9,10],\
                                 [5,6,7,8,9,10],\
                                 [5,6,7,8,9,10]],\
                             
                                [[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [5,6,7,8,9,10]],\
                             
                                [[5,6,7,8,9,10],\
                                 [5,6,7,8,9,10],\
                                 [0,1,2,3,4]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList7:#"前三直选单式","前三组选单式"
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01 02 03") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList6:#"前三组选复式"
                   clickList = [[[0,1,2,3,4]],\
                             
                                [[5,6,7,8,9,10]],\
                             
                                [[0,1,8,9,10]],\
                             
                                [[2,3,4,5,6,7]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList8:#前三组选胆拖"
                   clickList = [[[0,1],\
                                 [2,3,4,5,6,7,8,9,10]],\
                             
                                [[2,3],\
                                 [4,5,6,7,8,9,10,0,1]],\
                             
                                [[4,5],\
                                 [6,7,8,9,10,0,1,2,3]],\
                             
                                [[6,7],\
                                 [8,9,10,0,1,2,3,4,5]],\

                                [[8,9],\
                                 [10,0,1,2,3,4,5,6,7]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList1:#前三和值","中三和值","后三和值
                   for m in range(len(test_web.elements("div[class='syx5CheckNum fix'] ins" ,6))):
                       test_web.elementsClickOne("div[class='syx5CheckNum fix'] ins" ,6 ,m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

           elif test_web.webPlay()[j].text in ballList4: #二码
               if test_web.webPlayBranch()[k].text in ballBranchList5:#前二直选复式
                   clickList = [[[0,1,2,3,4],\
                                 [0,1,2,3,4]],\
                             
                                [[5,6,7,8,9,10],\
                                 [5,6,7,8,9,10]],\
                             
                                [[0,1,2,3,4],\
                                 [5,6,7,8,9,10]],\
                             
                                [[5,6,7,8,9,10],\
                                 [0,1,2,3,4]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList7:#"前二直选单式","前二组选单式"
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "01 02") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList6:#"前二组选复式"
                   clickList = [[[0,1,2,3,4]],\
                             
                                [[5,6,7,8,9,10]],\
                             
                                [[0,1,8,9,10]],\
                             
                                [[2,3,4,5,6,7]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList8:#前二组选胆拖"
                   clickList = [[[0],\
                                 [1,2,3,4,5,6,7,8,9,10]],\
                             
                                [[1],\
                                 [2,3,4,5,6,7,8,9,10,0]],\
                             
                                [[2],\
                                 [3,4,5,6,7,8,9,10,0,1]],\
                             
                                [[3],\
                                 [4,5,6,7,8,9,10,0,1,2]],\

                                [[4],\
                                 [5,6,7,8,9,10,0,1,2,3]],\

                                [[5],\
                                 [6,7,8,9,10,0,1,2,3,4]],\

                                [[6],\
                                 [7,8,9,10,0,1,2,3,4,5]],\
                                
                                [[7],\
                                 [8,9,10,0,1,2,3,4,5,6]],\

                                [[8],\
                                 [9,10,0,1,2,3,4,5,6,7]],\

                                [[9],\
                                 [10,0,1,2,3,4,5,6,7,8]],\

                                [[10],\
                                 [0,1,2,3,4,5,6,7,8,9]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList1:#头尾和值","前二和值","后二和值
                   for m in range(len(test_web.elements("div[class='syx5CheckNum fix'] ins" ,6))):
                       test_web.elementsClickOne("div[class='syx5CheckNum fix'] ins" ,6 ,m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")
              
sleep(waitSec)#等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_11選5" + "投注金額.xlsx")

if len(error) != 1:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_11選5_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()
