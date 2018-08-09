# -*- coding: utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#1.2修改單一投注數,增進效能,修改確認投注方式,判斷彩種是否開放

def submitCheck():
    period = []
    sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = test_web.webPage()[i].text
    sheet_money["C"+str(len(sheet_money["B"]))].value = Account               
    period.append(test_web.K3_r("input[type=text]" ,6 , max_Money = "1")) #投注金額

    sleep(1) #LHC專用

    try:
        period.insert(0 ,test_web.timeTitle()) #期號
    except:
        period.insert(0 ,"六合秒秒平常沒有期號") #期號
   
    submitCheck = True
    while(submitCheck):
        test_web.elementClick("div[class='Bet'] a[class='betBtn ClickShade UnClick']" ,6)
        #sleep(2)
        if test_web.radioWord() != "NO":#秒秒彩
            sleep(10)
            submitCheck = False
        if test_web.elementClick("div[class='section'] div[class='layermchild layerBet layermanim'] div[class='layermbtn'] span:nth-child(2)" ,6) != "NG": #六合彩不同之處
            #sleep(2)
            if test_web.submitCheckOK() != "NG":
                test_web.elementClick("//span[.='确定']" ,8)
                submitCheck = False
                
    #sleep(10)
   
    sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
    sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
    sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號
    
    sheet_row = len(sheet_money["B"]) #投注金額填表
    for k in range(len(period[1])):
        sheet_money.cell(row = sheet_row ,column = k + 7).value = period[1][k]
    wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_六合彩" + "投注金額.xlsx")


test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]

print("所有六合彩全餐投注")
wb = load_workbook("六合彩投注用.xlsx")
sheet = wb["六合彩"] # 獲取一張表
testNumber = input("測試站點序號:").strip()
for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        Account = str(sheet["F" + str(i)].value).strip()
        Password = str(sheet["G" + str(i)].value).strip()
        webPageSelect = str(sheet["I" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url) #目標網址
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']",6))
sleep(5)
test_web.webDriver.get(url) #目標網址

html_source = test_web.webDriver.page_source
if test_web.webDriver.current_url != url:
    input("此URL有誤,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "您所访问的彩种不存在，即将返回购彩大厅" in html_source:
    input("此彩種未開放,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "Unexpected token u in JSON at position 0" in html_source:
    input("此彩種未獎金模板錯誤或是HOST錯誤,請檢查,按enter離開。")
    test_web.webDriver.quit() 

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["六合彩"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

ballList = ["直选","任选","正１特","正２特","正３特","正４特","正５特","正６特"]
ballList1 = ["三全中","三中二","二全中","二中特","特串","五不中","六不中","七不中"]
ballList2 = ["八不中"]
ballList3 = ["九不中"]
ballList4 = ["十不中"]
boxList = ["两面","正１两面","正２两面","正３两面","正４两面","正５两面","正６两面",\
           "特码半波","特肖","一肖","总肖","正肖","特码头尾","七色波"]
boxList1 = ["二肖连","三肖连","四肖连","合肖中","合肖不中","二尾连","三尾连","四尾连"]
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH
for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
   if  webPageSelect != "1":
       test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁
        
   for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
       test_web.webPlayClick(j)
       for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
           test_web.webPlayBranchClick(k)
           #if test_web.webPlayBranch()[k].text in ballList or test_web.webPlayBranchLHC()[k].text in ballList:# ["直选","任选","正１特","正２特","正３特","正４特","正５特","正６特"]
           if test_web.webPlayBranch()[k].text in ballList:
               for m in range(len(test_web.webBall())):
                   test_web.webBallClick(m)
                   test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in ballList1:#["三全中","三中二","二全中","二中特","特串","五不中","六不中","七不中"]
               for m in range(len(test_web.webBall())):
                   test_web.webBallClick(m)
                   if (m+1)%7 == 0:
                       test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in ballList2:#["八不中"]
               for m in range(len(test_web.webBall())):
                   test_web.webBallClick(m)
                   if (m+1)%8 == 0:
                       test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in ballList3:#["九不中"]
               for m in range(len(test_web.webBall())):
                   test_web.webBallClick(m)
                   if (m+1)%9 == 0:
                       test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in ballList4:#["十不中"]
               for m in range(len(test_web.webBall())):
                   test_web.webBallClick(m)
                   if (m+1)%10 == 0:
                       test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in boxList:#["两面","正１两面","正２两面","正３两面","正４两面","正５两面","正６两面","特码半波","特肖","一肖","总肖","正肖","特码头尾","七色波"]
               test_web.elementsClickAll("ClickShade" ,2 ,len(test_web.elements("ClickShade" ,2))-1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in boxList1:#["二肖连","三肖连","四肖连","合肖中","合肖不中","二尾连","三尾连","四尾连"]
               for m in range(len(test_web.elements("ul[class='fix'] a[class='ClickShade']" ,6))):
                   test_web.elementsClickOne("ClickShade" ,2 ,m)
                   if (m+1)%4 == 0:
                       test_web.elementClick("betBtn" ,1)
               submitCheck()
           else:
               print(test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")
   
sleep(900) #等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_六合彩" + "投注金額.xlsx")

if len(error) != 1:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_六合彩_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()

