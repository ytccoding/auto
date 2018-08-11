# -*- coding: utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#1.3注數全拆開,修改確認投注方式,判斷彩種是否開放,等待時間,登入時間
print("快3全玩法投注")
testNumber = input("測試站點序號:").strip()
accountNumber = input("測試帳號序號:").strip()

test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]

wb = load_workbook("快3投注用.xlsx")
sheet = wb["快3"] # 獲取一張表
wbAccount = load_workbook("前台帳號.xlsx")
sheetAccount = wbAccount["帳號"] # 獲取一張表

for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        webPageSelect = str(sheet["G" + str(i)].value).strip()
        if str(sheet["F" + str(i)].value).strip() == "None":
            waitSec = 600
        else:
            waitSec = str(sheet["F" + str(i)].value).strip()

for i in range(1,len(sheetAccount["B"])+1):
    if str(sheetAccount["B" + str(i)].value).strip() == str(accountNumber):
        Account = str(sheetAccount["D" + str(i)].value).strip()
        Password = str(sheetAccount["E" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url)
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']" ,6))
#sleep(5)
timeCount = 0
while(test_web.webDriver.current_url != url):
    sleep(1)
    timeCount = timeCount + 1
    if timeCount >= 30:
        break
test_web.webDriver.get(url) #目標網址

html_source = test_web.webDriver.page_source
if test_web.webDriver.current_url != url:
    input("此URL有誤,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "您所访问的彩种不存在，即将返回购彩大厅" in html_source:
    input("此彩種未開放,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "Unexpected token u in JSON at position 0" in html_source:
    input("此彩種未獎金模板錯誤或是HOST錯誤,請檢查,按enter離開。")
    test_web.webDriver.quit()    
    
period = []

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["快3"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

boxList = ["和值","三同号通选","三连号通选"]
boxList2 = ["三同号单选","二同号复选"]
boxList3 = ["三不同号","二不同号"]
boxList4 = ["二同号单选"]
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH
    
for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
    if  webPageSelect != "1":
        test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁
    
    for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
        test_web.webPlayClick(j)
        if test_web.webPlay()[j].text in boxList:
            test_web.elementsClickAll("ClickShade" ,2 ,len(test_web.elements("a[class='ClickShade']",6)))
        elif test_web.webPlay()[j].text in boxList2:
            for k in range(len(test_web.elements("a[class='ClickShade']",6))):
                test_web.elementsClickOne("ClickShade" , 2 ,k)
                test_web.elementsClickOne("ClickShade" , 2 ,len(test_web.elements("a[class='ClickShade']",6))) 
        elif test_web.webPlay()[j].text in boxList3:
            clickList = [[0,1,2,3],\
                         [2,3,4,5],\
                         [0,1,4,5]] #可自由變換想投的內容,一律從0開始計算
            for k in range(len(clickList)):
                for m in range(len(clickList[k])):
                    test_web.elementsClickOne("ClickShade" , 2 ,clickList[k][m])
                test_web.elementsClickOne("ClickShade" , 2 ,len(test_web.elements("a[class='ClickShade']",6)))
        elif test_web.webPlay()[j].text in boxList4:
            clickList = [[0,1,2,9,10,11],\
                         [3,4,5,6,7,8],\
                         [2,3,6,7,10,11],\
                         [0,1,4,5,8,9]] #可自由變換想投的內容,一律從0開始計算
            for k in range(len(clickList)):
                for m in range(len(clickList[k])):
                    test_web.elementsClickOne("ClickShade" , 2 ,clickList[k][m])
                test_web.elementsClickOne("ClickShade" , 2 ,len(test_web.elements("a[class='ClickShade']",6)))
        else:
            print(test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

    sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = test_web.webPage()[i].text
    sheet_money["C"+str(len(sheet_money["B"]))].value = Account
    period.append(test_web.K3_r("input[type=text]" ,6 , max_Money = "1")) #投注金額

    try:
        period.insert(0 ,test_web.timeTitle()) #期號
    except:
        period.insert(0 ,"極速快3平常沒有期號") #期號
        
    submitCheck = True
    while(submitCheck):
        test_web.elementClick("a[class='betBtn ClickShade UnClick']" ,6)
        #sleep(2)
        if test_web.radioWord() == "YES": #秒秒彩
            sleep(10)
            submitCheck = False
        elif test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
            #sleep(2)
            if test_web.submitCheckOK() != "NG":
                test_web.elementClick("//span[.='确定']" ,8)
                submitCheck = False
    
    sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
    sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
    sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號
    
    sheet_row = len(sheet_money["B"]) #投注金額填表
    for k in range(len(period[1])):
        sheet_money.cell(row = sheet_row ,column = k + 7).value = period[1][k]
    wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_快3" + "投注金額.xlsx")
    period = []

sleep(waitSec) #等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_快3" + "投注金額.xlsx")

if len(error) != 1:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_快3_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()
