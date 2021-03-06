# -*- coding: utf-8 -*-

from selenium import webdriver
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#1.3移除不必要import,修改error輸出判斷,修改讀取的EXCEL
def submitCheck():
    period = []
    sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = "排列3" #排列3與其他彩種不同之處
    sheet_money["C"+str(len(sheet_money["B"]))].value = Account
    period.append(test_web.KL8("order_type" ,2)) #投注金額

    try:
        period.insert(0 ,test_web.timeTitle()) #期號
    except:
        period.insert(0 ,"秒秒彩平常沒有期號") #期號

    submitCheck = True
    while(submitCheck):
        test_web.elementClick("div[class='checkedListCon'] a[class='betBtn']" ,6)
        sleep(1)
        if test_web.radioWord() != "NO":
            sleep(10)
            submitCheck = False
        if test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
            if test_web.submitCheckOK() != "NG":
                test_web.elementClick("//span[.='确定']" ,8)
                submitCheck = False

    sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
    sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
    sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號
    
    sheet_row = len(sheet_money["B"]) #投注金額填表
    for k in range(len(period[1])):
        sheet_money.cell(row = sheet_row ,column = k + 7).value = period[1][k]
    wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_排列3" + "投注金額.xlsx")

print("排列3全玩法投注")
testNumber = input("測試站點序號:").strip()
accountNumber = input("測試帳號序號:").strip()

test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]

wb = load_workbook("投注URL.xlsx")
sheet = wb["排列3"] # 獲取一張表
wbAccount = load_workbook("前台帳號.xlsx")
sheetAccount = wbAccount["帳號"] # 獲取一張表

for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        webPageSelect = str(sheet["G" + str(i)].value).strip()
        if str(sheet["F" + str(i)].value).strip() == "None":
            waitSec = 600
        else:
            waitSec = int(str(sheet["F" + str(i)].value).strip())

for i in range(1,len(sheetAccount["B"])+1):
    if str(sheetAccount["B" + str(i)].value).strip() == str(accountNumber):
        Account = str(sheetAccount["D" + str(i)].value).strip()
        Password = str(sheetAccount["E" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url) #目標網址
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']",6))
timeCount = 0
while(test_web.webDriver.current_url != url):
    sleep(1)
    timeCount = timeCount + 1
    if timeCount >= 30:
        break
test_web.webDriver.get(url) #目標網址

html_source = test_web.webDriver.page_source
if test_web.webDriver.current_url != url:
    input("此URL有誤,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "您所访问的彩种不存在，即将返回购彩大厅" in html_source:
    input("此彩種未開放,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "Unexpected token u in JSON at position 0" in html_source:
    input("此彩種未獎金模板錯誤或是HOST錯誤,請檢查,按enter離開。")
    test_web.webDriver.quit() 

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["快樂彩"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

ballBranchList1 = ["直选和值","跨度","组选和值","组选包胆"]
ballBranchList2 = ["直选复式","复式"]
ballBranchList3 = ["直选单式","混合组选","组六单式","组选单式"]
ballBranchList4 = ["组三","组六","一码不定位","二码不定位","组选复式"]
ballBranchList5 = ["组三单式"]
ballBranchList6 = ["前二大小单双","后二大小单双"]
ballList1 = ["三星"]
ballList2 = ["前二","后二"]
ballList3 = ["一星"]
ballList4 = ["大小单双"]
boxList = []
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH
for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
   if  webPageSelect != "1":
       test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁
       
   for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
       test_web.webPlayClick(j)
       for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
           test_web.webPlayBranchClick(k)
           if test_web.webPlay()[j].text in ballList1: #三星
               if test_web.webPlayBranch()[k].text in ballBranchList1:#直选和值,跨度,组选和值,组选包胆
                   for m in range(len(test_web.webBall())):
                       test_web.webBallClick(m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList2:#三星直选复式
                   clickList = [[[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [0,1,2,3,4]],\
                             
                                [[5,6,7,8,9],\
                                 [5,6,7,8,9],\
                                 [5,6,7,8,9]],\
                             
                                [[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [5,6,7,8,9]],\
                             
                                [[5,6,7,8,9],\
                                 [5,6,7,8,9],\
                                 [0,1,2,3,4]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList3:#三星单式,組六单式
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "123") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList4:#三星组三
                   clickList = [[[0,1,2,3,4]],\
                             
                                [[5,6,7,8,9]],\
                             
                                [[0,1,7,8,9]],\
                             
                                [[2,3,4,5,6]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList5:#三星組三单式
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "001") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")
                   
           elif test_web.webPlay()[j].text in ballList2: #前二
               if test_web.webPlayBranch()[k].text in ballBranchList1:#直选和值,跨度,组选和值,组选包胆
                   for m in range(len(test_web.webBall())):
                       test_web.webBallClick(m)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList2:#前二直选复式
                   clickList = [[[0,1,2,3,4],\
                                 [0,1,2,3,4]],\
                             
                                [[5,6,7,8,9],\
                                 [5,6,7,8,9]],\
                             
                                [[0,1,2,3,4],\
                                 [5,6,7,8,9]],\
                             
                                [[5,6,7,8,9],\
                                 [0,1,2,3,4]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList3:#二星单式
                   test_web.elementClick("div[class='numberTextarea']" ,6)
                   test_web.elementSendKeys("textarea[class='betNote']" ,6 ,text = "12") #單式投注內容
                   sleep(1)
                   test_web.elementClick("betBtn" ,1)
                   submitCheck()
               elif test_web.webPlayBranch()[k].text in ballBranchList4:#二星组选复式
                   clickList = [[[0,1,2,3,4]],\
                             
                                [[5,6,7,8,9]],\
                             
                                [[0,1,7,8,9]],\
                             
                                [[2,3,4,5,6]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

           elif test_web.webPlay()[j].text in ballList3: #一星
               if  test_web.webPlayBranch()[k].text in ballBranchList2:#一星复式
                   clickList = [[[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [0,1,2,3,4]],\
                             
                                [[5,6,7,8,9],\
                                 [5,6,7,8,9],\
                                 [5,6,7,8,9]],\
                             
                                [[0,1,2,3,4],\
                                 [0,1,2,3,4],\
                                 [5,6,7,8,9]],\
                             
                                [[5,6,7,8,9],\
                                 [5,6,7,8,9],\
                                 [0,1,2,3,4]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")

           elif test_web.webPlay()[j].text in ballList4: #大小單雙
               if test_web.webPlayBranch()[k].text in ballBranchList6:#前二大小单双、后二大小单双
                   clickList = [[[0,1],\
                                 [0,1]],\
                             
                                [[2,3],\
                                 [2,3]],\
                             
                                [[0,1],\
                                 [2,3]],\
                             
                                [[2,3],\
                                 [0,1]]]
                   for m in range(len(clickList)):
                       for n in range(len(test_web.elements("div[class='buyNumber fix']" ,6))):
                           for o in range(len(clickList[m][n])):
                               test_web.webBallClick(clickList[m][n][o],n)
                       test_web.elementClick("betBtn" ,1)
                   submitCheck()
               else:
                   print(test_web.webPlay()[j].text + "_" + test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")
              
sleep(waitSec)#等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_排列3" + "投注金額.xlsx")

if len(error) != 1 and error[1] != None:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_排列3_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()
