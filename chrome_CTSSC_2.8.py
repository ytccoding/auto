# -*- coding: utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#2.8修改確認投注方式
test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]

def ctUse():
    global period
    try:
        period.insert(0 ,test_web.timeTitle()) #期號
    except:
        period.insert(0 ,"極速傳統時時彩平常沒有期號") #期號

    sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = test_web.webPage()[i].text
    sheet_money["C"+str(len(sheet_money["B"]))].value = Account

    sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
    sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
    sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號

    sheet_row = len(sheet_money["B"]) #投注金額填表
    for k in range(len(period[2])):
        sheet_money.cell(row = sheet_row ,column = k + 7).value = period[2][k]
    wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_傳統彩" + "投注金額.xlsx")
    period = []

print("所有傳統彩全餐投注")
wb = load_workbook("傳統彩投注用.xlsx")
sheet = wb["傳統彩"] # 獲取一張表
testNumber = input("測試站點序號:").strip()
for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        Account = str(sheet["F" + str(i)].value).strip()
        Password = str(sheet["G" + str(i)].value).strip()
        webPageSelect = str(sheet["I" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url) #目標網址
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']",6))
sleep(5)
test_web.webDriver.get(url) #目標網址

if test_web.webDriver.current_url != url:
    input("此彩種未開放或URL有誤,請檢查,按enter離開。")

period = []

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["傳統快3"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

chkBox = ["组选三","组选六"]
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH

for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
    if  webPageSelect != "1":
        test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁

    for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
        if j < 10:
            test_web.webPlayClick(j)
            for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
                test_web.webPlayBranchClick(k)
                period.append(test_web.webPlay()[j].text + ":" + test_web.webPlayBranch()[k].text)
                if test_web.webPlay()[j].text not in chkBox:
                    period.append(test_web.CTK3_r("input[type=text]" ,6 ,max_Money = "1")) #投注金額
                    ctUse()
                    submitCheck = True
                    while(submitCheck):
                        test_web.elementClick("button[class='btn btn-danger fl bet-add ']" ,6)
                        #sleep(2)
                        if test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
                            #sleep(2)
                            if test_web.submitCheckOK() != "NG":
                                test_web.elementClick("//span[.='确定']" ,8)
                                submitCheck = False
        elif j == 10:
            for m in range(8):
                test_web.morePlayClick(m)
                for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
                    test_web.webPlayBranchClick(k)
                    period.append(test_web.webPlay()[j].text + ":" + test_web.webPlayBranch()[k].text)
                    if test_web.webPlay()[j].text not in chkBox:
                        period.append(test_web.CTK3_r("input[type=text]" ,6 ,max_Money = "1")) #投注金額
                        ctUse()
                        submitCheck = True
                        while(submitCheck):
                            test_web.elementClick("button[class='btn btn-danger fl bet-add ']" ,6)
                            #sleep(2)
                            if test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
                                #sleep(2)
                                if test_web.submitCheckOK() != "NG":
                                    test_web.elementClick("//span[.='确定']" ,8)
                                    submitCheck = False
                  
sleep(600)#等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_傳統彩" + "投注金額.xlsx")

if len(error) != 1:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_傳統彩_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()

