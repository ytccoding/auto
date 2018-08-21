# -*- coding: utf-8 -*-

from selenium import webdriver
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#2.9移除不必要import,修改error輸出判斷,修改讀取的EXCEL

def submitCheck():
    period = []
    if test_web.webPlay()[j].text in chkBox:
        return "NG"
    
    for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
        test_web.webPlayBranchClick(k)
        period.append(test_web.webPlay()[j].text + ":" + test_web.webPlayBranch()[k].text)
        if test_web.webPlay()[j].text not in chkBox:
            period.append(test_web.CTK3_r("input[type=text]" ,6 ,max_Money = "1")) #投注金額

        sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = test_web.webPage()[i].text
        sheet_money["C"+str(len(sheet_money["B"]))].value = Account

        try:
            period.insert(0 ,test_web.timeTitle()) #期號
        except:
            period.insert(0 ,"極速傳統時時彩平常沒有期號") #期號

        submitCheck = True
        while(submitCheck):
            test_web.elementClick("button[class='btn btn-danger fl bet-add ']" ,6)
            sleep(1)
            if test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
                if test_web.submitCheckOK() != "NG":
                    test_web.elementClick("//span[.='确定']" ,8)
                    submitCheck = False

        sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
        sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
        sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號

        sheet_row = len(sheet_money["B"]) #投注金額填表
        for k in range(len(period[2])):
            sheet_money.cell(row = sheet_row ,column = k + 7).value = period[2][k]
        wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_傳統彩" + "投注金額.xlsx")

print("傳統彩全玩法投注")
testNumber = input("測試站點序號:").strip()
accountNumber = input("測試帳號序號:").strip()

test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]

wb = load_workbook("投注URL.xlsx")
sheet = wb["傳統彩"] # 獲取一張表
wbAccount = load_workbook("前台帳號.xlsx")
sheetAccount = wbAccount["帳號"] # 獲取一張表

for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        webPageSelect = str(sheet["G" + str(i)].value).strip()
        if str(sheet["F" + str(i)].value).strip() == "None":
            waitSec = 600
        else:
            waitSec = int(str(sheet["F" + str(i)].value).strip())

for i in range(1,len(sheetAccount["B"])+1):
    if str(sheetAccount["B" + str(i)].value).strip() == str(accountNumber):
        Account = str(sheetAccount["D" + str(i)].value).strip()
        Password = str(sheetAccount["E" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url) #目標網址
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']",6))

timeCount = 0
while(test_web.webDriver.current_url != url):
    sleep(1)
    timeCount = timeCount + 1
    if timeCount >= 30:
        break
test_web.webDriver.get(url) #目標網址

html_source = test_web.webDriver.page_source
if test_web.webDriver.current_url != url:
    input("此URL有誤,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "您所访问的彩种不存在，即将返回购彩大厅" in html_source:
    input("此彩種未開放,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "Unexpected token u in JSON at position 0" in html_source:
    input("此彩種未獎金模板錯誤或是HOST錯誤,請檢查,按enter離開。")
    test_web.webDriver.quit() 

period = []

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["傳統快3"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

chkBox = ["组选三","组选六"]
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH

for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
    if  webPageSelect != "1":
        test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁

    for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
        if j < 10:
            test_web.webPlayClick(j)
            submitCheck()
        elif j == 10:
            for m in range(8):
                test_web.morePlayClick(m)
                submitCheck()
                        
                  
sleep(waitSec)#等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_傳統彩" + "投注金額.xlsx")

if len(error) != 1 and error[1] != None:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_傳統彩_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()
