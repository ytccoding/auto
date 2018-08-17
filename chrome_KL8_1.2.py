# -*- coding: utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,random ,ytFuntion
#1.2修改單一投注數,增進效能,修改確認投注方式,判斷彩種是否開放,等待時間,登入時間
def submitCheck():
    period = []
    sheet_money["B"+str(len(sheet_money["B"]) + 1)].value = test_web.webPage()[i].text
    sheet_money["C"+str(len(sheet_money["B"]))].value = Account
   
    period.append(test_web.KL8("order_type" ,2)) #投注金額

    try:
        period.insert(0 ,test_web.timeTitle()) #期號
    except:
        period.insert(0 ,"秒秒平常沒有期號") #期號
   
    submitCheck = True
    while(submitCheck):
        test_web.elementClick("div[class='checkedListCon'] a[class='betBtn']" ,6)
        #sleep(2)
        if test_web.radioWord() == "YES":#秒秒彩
             sleep(10)
             submitCheck = False
        elif test_web.elementClick("//span[.='确认投注']" ,8) != "NG":
            #sleep(2)
            if test_web.submitCheckOK() != "NG":
                test_web.elementClick("//span[.='确定']" ,8)
                submitCheck = False
           
    #sleep(10)
   
    sheet_money["D"+str(len(sheet_money["B"]))].value = time.strftime("%y_%m_%d") #投注時間
    sheet_money["E"+str(len(sheet_money["B"]))].value = time.strftime("%H_%M_%S") #投注時間
    sheet_money["F"+str(len(sheet_money["B"]))].value = period[0] #投注期號
    
    sheet_row = len(sheet_money["B"]) #投注金額填表
    for k in range(len(period[1])):
        sheet_money.cell(row = sheet_row ,column = k + 7).value = period[1][k]
    wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_快樂8" + "投注金額.xlsx")

print("快樂8全玩法投注")
testNumber = input("測試站點序號:").strip()
accountNumber = input("測試帳號序號:").strip()

test_web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe'))
error = ["ERROR:"]


wb = load_workbook("快樂8投注用.xlsx")
sheet = wb["快樂彩"] # 獲取一張表
wbAccount = load_workbook("前台帳號.xlsx")
sheetAccount = wbAccount["帳號"] # 獲取一張表

for i in range(1,len(sheet["B"])+1):
    if str(sheet["B" + str(i)].value).strip() == str(testNumber):
        url = str(sheet["E" + str(i)].value).strip()
        webPageSelect = str(sheet["G" + str(i)].value).strip()
        if str(sheet["F" + str(i)].value).strip() == "None":
            waitSec = 600
        else:
            waitSec = int(str(sheet["F" + str(i)].value).strip())

for i in range(1,len(sheetAccount["B"])+1):
    if str(sheetAccount["B" + str(i)].value).strip() == str(accountNumber):
        Account = str(sheetAccount["D" + str(i)].value).strip()
        Password = str(sheetAccount["E" + str(i)].value).strip()
        
print("使用帳號:" + Account)

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

test_web.webDriver.get(url) #目標網址
test_web.webDriver.maximize_window()

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

error.append(test_web.elementClick("亲，请登录",3))
test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = Account)
test_web.elementSendKeys("input[tag=密码]" ,6 ,text = Password)

error.append(test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']",6))
#sleep(5)
timeCount = 0
while(test_web.webDriver.current_url != url):
    sleep(1)
    timeCount = timeCount + 1
    if timeCount >= 30:
        break
test_web.webDriver.get(url) #目標網址

html_source = test_web.webDriver.page_source
if test_web.webDriver.current_url != url:
    input("此URL有誤,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "您所访问的彩种不存在，即将返回购彩大厅" in html_source:
    input("此彩種未開放,請檢查,按enter離開。")
    test_web.webDriver.quit()
elif "Unexpected token u in JSON at position 0" in html_source:
    input("此彩種未獎金模板錯誤或是HOST錯誤,請檢查,按enter離開。")
    test_web.webDriver.quit() 
    
period = []

wb_money = load_workbook("投注金額.xlsx")
sheet_money = wb_money["快樂彩"] # 獲取金額表
sheet_money["D1"].value = url
test_web.showMoneyClick()
sheet_money["H1"].value = test_web.getMoney() #投注前金額

ballList = ["任选一","任选二","任选三","任选四","任选五","任选六","任选七"]
boxList = ["上下盘","奇偶盘","和值大小单双"]
#ballNum = {"任选一":"1","任选二":"2","任选三":"3","任选四":"4","任选五":"5","任选六":"6","任选七":"7"}
#全部全餐

#1=ID,2=CLASS_NAME,3=LINK_TEXT,4=PARTIAL_LINK_TEXT,5=NAME,6=CSS_SELECTOR,7=TAG_NAME,8=XPATH
for i in range(test_web.webPageSelect(webPageSelect)): #所有分頁
   if  webPageSelect != "1":
       test_web.webPageClick(i ,"a[class ='betNavtab right']" ,6) #切換分頁
       
   for j in range(len(test_web.webPlay())): #該分頁所有可點選玩法都點
       test_web.webPlayClick(j)
       for k in range(len(test_web.webPlayBranch())):#該分頁所有可點選玩法分支都點
           test_web.webPlayBranchClick(k)
           if test_web.webPlayBranch()[k].text in ballList:
               for m in range(len(test_web.webBall(0)) + len(test_web.webBall(1))):
                   if m < 40: #上下盤切換
                       test_web.webBallClick(m ,0)
                   else:
                       test_web.webBallClick(m%40 ,1)
                   if test_web.webPlayBranch()[k].text == "任选一":
                       test_web.elementClick("betBtn" ,1)
                   elif (m+1)%8 == 0: #8個一注
                       test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in boxList[0]:
               for k in range(1 ,4):
                   test_web.elementsClickOne("ins" , 7 ,k)
                   test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in boxList[1]:
               for k in range(1 ,4):
                   test_web.elementsClickOne("ins" , 7 ,k)
                   test_web.elementClick("betBtn" ,1)
               submitCheck()
           elif test_web.webPlayBranch()[k].text in boxList[2]:
               for k in range(3 ,7):
                   test_web.elementsClickOne("ins" , 7 ,k)
                   test_web.elementClick("betBtn" ,1)
               submitCheck()
           else:
               print(test_web.webPlayBranch()[k].text + "玩法自動投注尚未完成")
       


sleep(waitSec) #等全部開獎完畢
test_web.elementClick("更多>>" ,3) #投注明細

periodDetail = test_web.periodDetail()
sheet_detail = ytFuntion.sheet_work(wb_money["投注紀錄"]) # 獲取金額表
sheet_detail.periodDetail(periodDetail)

test_web.reflashMoney()
sheet_detail.sheet_work["R2"].value = test_web.getMoney() #投注後含派彩金額
wb_money.save(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_快樂8" + "投注金額.xlsx")

if len(error) != 1:
    Error = open(os.getcwd() + "\\" + str(testdayFile) + "\\" + str(testdayTime) + "_快樂8_ERROR.txt" ,"wb+")
    for i in error:
        Error.write(str(i).encode('utf-8'))
    Error.close()

test_web.webDriver.quit()
